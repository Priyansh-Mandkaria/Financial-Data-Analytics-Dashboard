"""
Excel Report Generator
Creates formatted Excel workbooks with multiple sheets.
"""

import os
import sys
import pandas as pd
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database.db_manager import execute_query


def generate_excel_report(output_dir: str = None) -> str:
    """
    Generate a comprehensive Excel report from the database.
    Returns the path to the generated file.
    """
    if output_dir is None:
        output_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'reports', 'output'
        )
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = os.path.join(output_dir, f'Financial_Analytics_Report_{timestamp}.xlsx')

    print("\n  [REPORT] Generating Excel report...")

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # ── Sheet 1: Executive Summary ──
        summary_data = _build_summary_data()
        summary_data.to_excel(writer, sheet_name='Executive Summary', index=False)
        _format_sheet(writer, 'Executive Summary', summary_data)

        # ── Sheet 2: Transaction Details ──
        txn_df = execute_query("""
            SELECT transaction_id, transaction_date, account_id, category,
                   subcategory, amount, currency, transaction_type, status, counterparty
            FROM transactions ORDER BY transaction_date DESC LIMIT 5000
        """)
        txn_df.to_excel(writer, sheet_name='Transactions', index=False)
        _format_sheet(writer, 'Transactions', txn_df)

        # ── Sheet 3: Quality Report ──
        quality_df = execute_query("""
            SELECT check_name, check_type, column_name, total_records,
                   failed_records, pass_rate, severity, details
            FROM quality_log ORDER BY pass_rate ASC
        """)
        quality_df.to_excel(writer, sheet_name='Quality Report', index=False)
        _format_sheet(writer, 'Quality Report', quality_df, highlight_col='severity')

        # ── Sheet 4: Anomalies ──
        anomaly_df = execute_query("""
            SELECT transaction_id, anomaly_type, anomaly_score,
                   expected_range_low, expected_range_high, actual_value,
                   category, detection_method
            FROM anomaly_log ORDER BY anomaly_score DESC LIMIT 2000
        """)
        anomaly_df.to_excel(writer, sheet_name='Anomalies', index=False)
        _format_sheet(writer, 'Anomalies', anomaly_df)

        # ── Sheet 5: Reconciliation Summary ──
        recon_df = _build_reconciliation_summary()
        recon_df.to_excel(writer, sheet_name='Reconciliation', index=False)
        _format_sheet(writer, 'Reconciliation', recon_df)

        # ── Sheet 6: Benchmark Tracking ──
        bench_df = execute_query("""
            SELECT benchmark_date, benchmark_name, expected_value,
                   actual_value, variance, variance_pct, status
            FROM benchmarks ORDER BY benchmark_date DESC
        """)
        bench_df.to_excel(writer, sheet_name='Benchmarks', index=False)
        _format_sheet(writer, 'Benchmarks', bench_df, highlight_col='status')

    print(f"  ✓ Report saved: {filepath}")
    return filepath


def _build_summary_data() -> pd.DataFrame:
    """Build executive summary metrics."""
    try:
        txn = execute_query("SELECT COUNT(*) as cnt, SUM(amount) as total, AVG(amount) as avg FROM transactions")
        quality = execute_query("SELECT AVG(pass_rate) as score FROM quality_log")
        anomalies = execute_query("SELECT COUNT(*) as cnt FROM anomaly_log")
        completed = execute_query("SELECT COUNT(*) as cnt FROM transactions WHERE status='completed'")

        total_count = int(txn['cnt'].iloc[0])
        metrics = {
            'Metric': [
                'Total Records Processed',
                'Total Transaction Volume (USD)',
                'Average Transaction Amount',
                'Transaction Completion Rate',
                'Data Quality Score',
                'Total Anomalies Detected',
                'Report Generated',
            ],
            'Value': [
                f"{total_count:,}",
                f"${float(txn['total'].iloc[0]):,.2f}",
                f"${float(txn['avg'].iloc[0]):,.2f}",
                f"{completed['cnt'].iloc[0] / total_count * 100:.1f}%" if total_count > 0 else 'N/A',
                f"{float(quality['score'].iloc[0]):.1f}%" if pd.notna(quality['score'].iloc[0]) else 'N/A',
                f"{int(anomalies['cnt'].iloc[0]):,}",
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            ],
        }
        return pd.DataFrame(metrics)
    except Exception:
        return pd.DataFrame({'Metric': ['Error'], 'Value': ['Could not load data']})


def _build_reconciliation_summary() -> pd.DataFrame:
    """Build reconciliation summary data."""
    try:
        txn_count = int(execute_query("SELECT COUNT(*) as cnt FROM transactions WHERE status='completed'")['cnt'].iloc[0])
        led_count = int(execute_query("SELECT COUNT(*) as cnt FROM ledger_entries")['cnt'].iloc[0])
        matched = int(execute_query("""
            SELECT COUNT(*) as cnt FROM transactions t
            INNER JOIN ledger_entries l ON t.transaction_id = l.transaction_id
            WHERE t.status='completed'
        """)['cnt'].iloc[0])
        mismatches = int(execute_query("""
            SELECT COUNT(*) as cnt FROM transactions t
            INNER JOIN ledger_entries l ON t.transaction_id = l.transaction_id
            WHERE t.status='completed'
            AND ABS(t.amount - CASE WHEN l.debit_amount > 0 THEN l.debit_amount ELSE l.credit_amount END) > 0.01
        """)['cnt'].iloc[0])

        metrics = {
            'Metric': [
                'Completed Transactions',
                'Ledger Entries',
                'Matched Records',
                'Match Rate',
                'Amount Mismatches',
                'Unmatched Transactions',
                'Perfect Matches',
            ],
            'Value': [
                f'{txn_count:,}', f'{led_count:,}', f'{matched:,}',
                f'{matched/txn_count*100:.1f}%' if txn_count > 0 else 'N/A',
                f'{mismatches:,}', f'{txn_count - matched:,}', f'{matched - mismatches:,}',
            ],
        }
        return pd.DataFrame(metrics)
    except Exception:
        return pd.DataFrame({'Metric': ['Error'], 'Value': ['Could not load data']})


def _format_sheet(writer, sheet_name: str, df: pd.DataFrame, highlight_col: str = None):
    """Apply formatting to an Excel sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]

    # Header styling
    header_fill = PatternFill(start_color='1a1b2e', end_color='1a1b2e', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='8b9dc3')
    thin_border = Border(
        bottom=Side(style='thin', color='2d3748'),
    )

    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

    # Auto-adjust column widths
    for col_idx, col in enumerate(df.columns, 1):
        max_length = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) > 0 else 0
        )
        adjusted_width = min(max_length + 4, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Conditional formatting for severity / status columns
    if highlight_col and highlight_col in df.columns:
        col_idx = df.columns.tolist().index(highlight_col) + 1
        red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        amber_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')

        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = str(cell.value).lower()
            if val in ('critical', 'high'):
                cell.fill = red_fill
                cell.font = Font(color='DC2626', bold=True)
            elif val in ('medium', 'warning'):
                cell.fill = amber_fill
                cell.font = Font(color='D97706')
            elif val in ('low', 'within_range'):
                cell.fill = green_fill
                cell.font = Font(color='059669')


if __name__ == '__main__':
    generate_excel_report()
